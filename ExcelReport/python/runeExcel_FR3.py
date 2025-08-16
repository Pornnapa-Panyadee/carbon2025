import os
import sys
import shutil
from datetime import datetime
import subprocess
import platform
from openpyxl.styles import PatternFill,Font,Alignment
from openpyxl.styles import Border, Side

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import requests
import json

import sys
import os
os.system("chcp 65001")  # เปลี่ยน code page เป็น UTF-8
sys.stdout.reconfigure(encoding='utf-8')  # Python 3.7+

# ----------------- Arguments -----------------
if len(sys.argv) < 3:
    print("Usage: python runExcel.py <company_name> <product_1>")
    sys.exit(1)

company_name = sys.argv[1]
product_1 = sys.argv[2]

# ----------------- API Call -----------------
url = f"http://localhost:5000/api/v1/f1/excel/{company_name}/{product_1}"
resp = requests.get(url)
if resp.status_code != 200:
    print("เกิดข้อผิดพลาดในการเรียก API:", resp.status_code)
    sys.exit(1)

data = resp.json()
product = data.get("product", {})
company = data.get("company", {})
process = data.get("process", [])

# ----------------- Helper -----------------
def thai_date_format(iso_date_str):
    if not iso_date_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_date_str.replace("Z", "+00:00"))
    except ValueError:
        return "รูปแบบวันที่ไม่ถูกต้อง"

    months_th = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม",
                 "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม",
                 "พฤศจิกายน", "ธันวาคม"]
    return f"{dt.day} {months_th[dt.month]} {dt.year + 543}"

# ----------------- Prepare Data -----------------
start = thai_date_format(product.get("collect_data_start"))
end = thai_date_format(product.get("collect_data_end"))
date_range = f"{start} - {end}"
submitted_date_thai = thai_date_format(product.get("submitted_date"))

techinfo_raw = product.get("product_techinfo") or "[]"
techinfo_list = json.loads(techinfo_raw) if techinfo_raw else []
product_techinfo_array = [str(i).strip() for i in techinfo_list]

# ----------------- File Paths -----------------
template_path = "ExcelReport/excel/form_CFP_pdf.xlsx"
timestamp = datetime.now().strftime("%Y")
output_dir = "ExcelReport/output"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

output_xlsx = os.path.join(output_dir, f"{timestamp}_Company{company.get('company_id','')}_Product{product.get('product_id','')}.xlsx")
shutil.copy(template_path, output_xlsx)

# ----------------- Load Workbook -----------------

def set_borderThinTop(ws, cell_range,position):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            if position == 'left':
                cell.border = Border(left=Side(style='thin'))
            elif position == 'right':
                cell.border = Border(right=Side(style='thin'))
            elif position == 'top':
                cell.border = Border(
                    top=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )
            elif position == 'bottom':
                 cell.border = Border(
                    bottom=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )
    



######## ---------------------------------------------------------------------
wb = load_workbook(output_xlsx)
ws03 = wb["Fr-03"]   
ws03["J1"] = date_range
ws03["D2"] = company.get("name", "")
ws03["D3"] = product.get("product_name_th", "")
fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid') 
fill_intput_head = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid')
fill_intput = PatternFill(start_color='FFCCFFCC', end_color='FFCCFFCC', fill_type='solid')
fill_output = PatternFill(start_color='FFFFCC99', end_color='FFFFCC99', fill_type='solid')
fill_waste = PatternFill(start_color='FFCCFFFF', end_color='FFCCFFFF', fill_type='solid')
fill_arrow = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
fill_tail_title = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
fill_light_blue = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')  # ฟ้าอ่อน
font_white = Font(color="FFFFFF", bold=True, size=14)
align_center = Alignment(horizontal="center", vertical="center")

row = 13 
for i in range(len(process)):
    input_names = []
    row_input = row
    row_input_offset = 0
    max_input_rows = 0 

    # ===== Input Section =====
    input_categories = process[i]['inputs']
    num_categories = len(input_categories)

    for j in range(max(1, num_categories)):
        if num_categories > 0:
            category = input_categories[j]
            items = category['items']
            num_items = len(items)
            input_title = category['input_title']
        else:
            items = []
            num_items = 0
            input_title = "ไม่มีข้อมูล"  # หรือจะเป็น "" ตามที่คุณต้องการ

        current_row = row_input + row_input_offset
        ws03.merge_cells(f"B{current_row}:D{current_row}")
        ws03[f"B{current_row}"] = input_title
        ws03[f"B{current_row}"].fill = fill_intput_head

        # ถ้ามี item จริง
        if num_items > 0:
            for k in range(num_items):
                ws03[f"B{current_row + k + 1}"] = items[k]['input_name']
                ws03[f"C{current_row + k + 1}"] = items[k]['input_quantity']
                ws03[f"D{current_row + k + 1}"] = items[k]['input_unit']

                ws03[f"B{current_row + k + 1}"].fill = fill_intput
                ws03[f"C{current_row + k + 1}"].fill = fill_intput
                ws03[f"D{current_row + k + 1}"].fill = fill_intput
        else:
            # ใส่ช่องว่างแถวล่างไว้ 1 แถวพร้อมสี
            ws03[f"B{current_row + 1}"] = ""
            ws03[f"C{current_row + 1}"] = ""
            ws03[f"D{current_row + 1}"] = ""

            ws03[f"B{current_row + 1}"].fill = fill_intput
            ws03[f"C{current_row + 1}"].fill = fill_intput
            ws03[f"D{current_row + 1}"].fill = fill_intput

        row_input_offset += 1 + max(1, num_items)
        max_input_rows += 1 + max(1, num_items)
    # ===== Process Block ด้านขวา =====
    process_row = row
    ws03.merge_cells(f"F{process_row}:I{process_row}")
    ws03[f"F{process_row}"] = process[i]['ordering']
    ws03[f"F{process_row}"].fill = fill
    set_borderThinTop(ws03, f"F{process_row}:I{process_row}", 'top')

    ws03.merge_cells(f"F{process_row + 1}:I{process_row + 5}")
    ws03[f"F{process_row + 1}"] = process[i]['process_name']
    ws03[f"F{process_row + 1}"].fill = fill
    set_borderThinTop(ws03, f"F{process_row + 1}:I{process_row + 5}", 'bottom')

    # ===== output =====
    ws03[f"I{process_row + 7}"] = "ผลิตภัณฑ์"
    ws03[f"I{process_row + 7}"].fill = fill_intput_head 

    for j in range(len(process[i]['outputs'])):
        output = process[i]['outputs'][j]
        ws03[f"I{process_row + 8 + j}"] = output['output_name']
        ws03[f"J{process_row + 8 + j}"] = output['output_quantity']
        ws03[f"K{process_row + 8 + j}"] = output['output_unit']
        ws03[f"I{process_row + 8 + j}"].fill = fill_output
        ws03[f"J{process_row + 8 + j}"].fill = fill_output
        ws03[f"K{process_row + 8 + j}"].fill = fill_output

    # ===== waste =====
    waste_head=["ผลิตภัณฑ์ร่วม", "ของเสีย"]
    row_input_offset = 0
    current_row = row_input + row_input_offset
    waste_total_rows = 0 

    if process[i]['wastes'] == []:
        ws03.merge_cells(f"K{current_row}:M{current_row}")
        ws03[f"K{current_row}"] = "ผลิตภัณฑ์ร่วม"
        ws03[f"K{current_row}"].fill = fill_intput_head
        ws03.merge_cells(f"K{current_row + 1}:M{current_row + 1}")
        ws03[f"K{current_row + 1}"].fill = fill_waste

        ws03.merge_cells(f"K{current_row+2}:M{current_row+2}")
        ws03[f"K{current_row+2}"] = "ของเสีย"
        ws03[f"K{current_row+2}"].fill = fill_intput_head
        ws03.merge_cells(f"K{current_row + 3}:M{current_row + 3}")
        ws03[f"K{current_row + 3}"].fill = fill_waste
    else:
        len_waste = len(process[i]['wastes'])
        if(len_waste == 1):
            if process[i]['wastes'][0]['waste_cat_name'] == waste_head[0]:
                ws03.merge_cells(f"K{current_row}:M{current_row}")
                ws03[f"K{current_row}"] = waste_head[0]
                ws03[f"K{current_row}"].fill = fill_intput_head
                for j in range(len(process[i]['wastes'][0]['items'])):
                    item = process[i]['wastes'][0]['items'][j]
                    ws03[f"K{current_row + 2 + j}"] = item['waste_name']
                    ws03[f"L{current_row + 2 + j}"] = item['waste_qty']
                    ws03[f"M{current_row + 2 + j}"] = item['waste_unit']
                    ws03[f"K{current_row + 2 + j}"].fill = fill_waste
                    ws03[f"L{current_row + 2 + j}"].fill = fill_waste
                    ws03[f"M{current_row + 2 + j}"].fill = fill_waste
                
                
                waste_total_rows += len(process[i]['wastes'][0]['items']) + 2

                ws03.merge_cells(f"K{waste_total_rows}:M{waste_total_rows}")
                ws03[f"K{waste_total_rows}"] = waste_head[0]
                ws03[f"K{waste_total_rows}"].fill = fill_intput_head
                ws03.merge_cells(f"K{waste_total_rows+1}:M{waste_total_rows+1}")
                ws03[f"K{waste_total_rows+1}"].fill = fill_waste

            else:

                ws03.merge_cells(f"K{current_row}:M{current_row}")
                ws03[f"K{current_row}"] = waste_head[0]
                ws03[f"K{current_row}"].fill = fill_intput_head
                ws03.merge_cells(f"K{current_row + 1}:M{current_row + 1}")
                ws03[f"K{current_row + 1}"].fill = fill_waste
                ws03.merge_cells(f"K{current_row + 2}:M{current_row + 2}")
                ws03[f"K{current_row + 2}"] = waste_head[1] 
                ws03[f"K{current_row + 2}"].fill = fill_intput_head


                for j in range(len(process[i]['wastes'][0]['items'])):
                    item = process[i]['wastes'][0]['items'][j]
                    ws03[f"K{current_row + 3 + j}"] = item['waste_name']
                    ws03[f"L{current_row + 3 + j}"] = item['waste_qty']
                    ws03[f"M{current_row + 3 + j}"] = item['waste_unit']
                    ws03[f"K{current_row + 3 + j}"].fill = fill_waste
                    ws03[f"L{current_row + 3 + j}"].fill = fill_waste
                    ws03[f"M{current_row + 3 + j}"].fill = fill_waste

                waste_total_rows += len(process[i]['wastes'][0]['items']) + 2

        for j in range(len(process[i]['wastes'])):

            if process[i]['wastes'][0]['waste_cat_name']==waste_head[j]:
                ws03.merge_cells(f"K{current_row}:M{current_row}")
                ws03[f"K{current_row}"] = waste_head[0]
                ws03[f"K{current_row}"].fill = fill_intput_head
                 # เขียนรายการผลิตภัณฑ์ร่วม
                for j in range(len(process[i]['wastes'][0]['items'])):
                    item = process[i]['wastes'][0]['items'][j]
                    ws03[f"K{current_row + 2 + j}"] = item['waste_name']
                    ws03[f"L{current_row + 2 + j}"] = item['waste_qty']
                    ws03[f"M{current_row + 2 + j}"] = item['waste_unit']
                    ws03[f"K{current_row + 2 + j}"].fill = fill_waste
                    ws03[f"L{current_row + 2 + j}"].fill = fill_waste
                    ws03[f"M{current_row + 2 + j}"].fill = fill_waste

                waste_total_rows += len(process[i]['wastes'][0]['items']) + 2
            else:
                ws03.merge_cells(f"K{current_row}:M{current_row}")
                ws03[f"K{current_row}"] = waste_head[0]
                ws03[f"K{current_row}"].fill = fill_intput_head
                ws03.merge_cells(f"K{current_row + 1}:M{current_row + 1}")
            
            ws03[f"K{current_row + 1}"].fill = fill_waste

    # ===== ปรับ row สำหรับ process ถัดไป =====
    ws03[f"E{process_row+2}"].fill = fill_arrow
    ws03[f"J{process_row+2}"].fill = fill_arrow
    # ws03.merge_cells(f"H{process_row+6}:H{process_row+6+4}")
    # ws03[f"H{process_row+6}"].fill = fill_arrow

    if (max_input_rows <=6):
        u=process_row+6
        l=process_row+9
    else:
        x=max_input_rows-9
        u=process_row+6
        l=process_row+10+x
    
    if j< len(process)-1:
        ws03.merge_cells(f"G{u}:G{l}")
        ws03[f"G{u}"].fill = fill_arrow

    row += max(max_input_rows + 2, 10) 


# กำหนดค่าหลัก
ws03[f"A{row+5}"] = "จัดทำโดย"
ws03[f"E{row+5}"] = "เสร็จสิ้นวันที่"
ws03[f"J{row+5}"] = "วันที่แก้ไข"

# Merge E-F
ws03.merge_cells(f"E{row+5}:F{row+5}")
ws03[f"E{row+5}"].alignment = align_center

# ใส่ fill น้ำเงิน + font ขาว
for col in ["A", "E", "J"]:
    cell = ws03[f"{col}{row+5}"]
    cell.fill = fill_tail_title
    cell.font = font_white

# ใส่ฟ้าอ่อนใน B-D, G-I, K-M
for col_range in ["B:D", "G:I", "K:M"]:
    start_col, end_col = col_range.split(":")
    cell_range = f"{start_col}{row+5}:{end_col}{row+5}"
    for row_cells in ws03[cell_range]:
        for cell in row_cells:
            cell.fill = fill_light_blue
######## ---------------------------------------------------------------------

wb.save(output_xlsx)
print("Excel saved:", output_xlsx)

# ----------------- Export Sheet Fr-01 to PDF -----------------
def export_pdf_windows_excel(input_file, output_file):
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb_com = excel.Workbooks.Open(os.path.abspath(input_file))
    ws_com = wb_com.Sheets("Fr-03")
    ws_com.ExportAsFixedFormat(0, os.path.abspath(output_file))
    wb_com.Close(False)
    excel.Quit()

def export_pdf_libreoffice(input_file, output_dir):
    # ต้องระบุ path ของ LibreOffice หรือ soffice.exe
    if platform.system() == "Windows":
        libreoffice_path = r"C:\Program Files\LibreOffice\program\soffice.exe"
    else:
        libreoffice_path = "/usr/bin/libreoffice"
    if not os.path.exists(libreoffice_path):
        raise FileNotFoundError(f"LibreOffice not found at {libreoffice_path}")

    subprocess.run([
        libreoffice_path,
        "--headless",
        "--convert-to", "pdf",
        "--outdir", output_dir,
        input_file
    ], check=True)

# PDF output path
output_pdf = os.path.join(output_dir, f"{timestamp}_Company{company.get('company_id','')}_Product{product.get('product_id','')}_Fr03.pdf")

try:
    if platform.system() == "Windows":
        try:
            export_pdf_windows_excel(output_xlsx, output_pdf)
            print("PDF saved via Excel COM:", output_pdf)
        except Exception:
            # ถ้าไม่มี Excel ใช้ LibreOffice
            export_pdf_libreoffice(output_xlsx, output_pdf)
            print("PDF saved via LibreOffice:", output_pdf)
    else:
        export_pdf_libreoffice(output_xlsx, output_pdf)
        print("PDF saved via LibreOffice:", output_pdf)
except Exception as e:
    print("PDF export failed:", e)
