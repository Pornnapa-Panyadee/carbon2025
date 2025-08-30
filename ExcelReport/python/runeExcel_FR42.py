import os
import sys
import shutil
from datetime import datetime
import subprocess
import platform
from openpyxl.styles import PatternFill,Font,Alignment
from openpyxl.styles import Border, Side

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import requests
import json
import re

os.system("chcp 65001")  # เปลี่ยน code page เป็น UTF-8
sys.stdout.reconfigure(encoding='utf-8')  # Python 3.7+

# ----------------- Arguments -----------------
if len(sys.argv) < 3:
    print("Usage: python runExcel.py <company_name> <product_1>")
    sys.exit(1)
company_name = sys.argv[1]
product_1 = sys.argv[2]

# ----------------- API Call -----------------
url = f"http://localhost:5000/api/v1/f4-2/form/{product_1}"
resp = requests.get(url)
if resp.status_code != 200:
    print("เกิดข้อผิดพลาดในการเรียก API:", resp.status_code)
    sys.exit(1)

data = resp.json()

form42 = data.get("form42", {})
company = data.get("company", {})
product = data.get("product", {})
process = data.get("process", {})
report42Sum = data.get("report42Sum", [])[0]




# ----------------- Helper -----------------
def thai_date_format(iso_date_str):
    if not iso_date_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_date_str.replace("Z", "+00:00"))
    except ValueError:
        return "รูปแบบวันที่ไม่ถูกต้อง"

    months_th = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม",
                 "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม",
                 "พฤศจิกายน", "ธันวาคม"]
    return f"{dt.day} {months_th[dt.month]} {dt.year + 543}"

def check_source_form4_2_T1(type1_ef_source, row, ws_name='ws42'):
    ef_map = { 'TGO EF': 'K','Int. DB': 'L'}

    if type1_ef_source in ef_map:
        col = ef_map[type1_ef_source]
        cell = f'{col}{row}'
        return f'{ws_name}["{cell}"] = "●"'
    else:
        return f'# type1_ef_source \"{type1_ef_source}\" not recognized'

def check_source_form4_2_T2(type2_ef_source, row, ws_name='ws42'):
    ef_map = { 'TGO EF': 'T','Int. DB': 'U'}

    if type2_ef_source in ef_map:
        col = ef_map[type2_ef_source]
        cell = f'{col}{row}'
        return f'{ws_name}["{cell}"] = "●"'
    else:
        return f'# type2_ef_source \"{type2_ef_source}\" not recognized'


    

# ----------------- Prepare Data -----------------
start = thai_date_format(product.get("collect_data_start"))
end = thai_date_format(product.get("collect_data_end"))
date_range = f"{start} - {end}"
submitted_date_thai = thai_date_format(product.get("submitted_date"))

techinfo_raw = product.get("product_techinfo") or "[]"
techinfo_list = json.loads(techinfo_raw) if techinfo_raw else []
product_techinfo_array = [str(i).strip() for i in techinfo_list]

# ----------------- File Paths -----------------
template_path = "ExcelReport/excel/form_CFP_pdf042.xlsx"
timestamp = datetime.now().strftime("%Y")
output_dir = "ExcelReport/output"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

output_xlsx = os.path.join(output_dir, f"{timestamp}_Company{company.get('company_id','')}_Product{product.get('product_id','')}.xlsx")
shutil.copy(template_path, output_xlsx)

# ----------------- Load Workbook -----------------

def set_borderThinTop(ws, cell_range,position):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            if position == 'left':
                cell.border = Border(left=Side(style='thin'))
            elif position == 'right':
                cell.border = Border(right=Side(style='thin'))
            elif position == 'top':
                cell.border = Border(
                    top=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )
            elif position == 'bottom':
                 cell.border = Border(
                    bottom=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )

######## ---------------------------------------------------------------------
wb = load_workbook(output_xlsx)
ws42 = wb["Fr-04.2"]
fill_sum = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid') 
r_start = 12
row = 12
phase = ["การได้มาของวัตถุดิบ", "การผลิต", "การกระจายสินค้า", "การใช้งาน", "การจัดการซาก"]
for i in range(len(phase)):
    phase_start_row = row  # เก็บแถวเริ่มต้นของ phase นี้
    sum_proportion = 0  # กำหนดค่าเริ่มต้นของ sum_proportion

    for j in range(len(form42[i]["process"])):
        process = form42[i]["process"][j]
        process_start_row = row  # เก็บแถวเริ่มต้นของ process นี้

        if i < 2:
            ws42[f"B{row}"] = process["process_name"]
            ws42[f"B{row}"].font = Font(name="Tahoma", size=10, bold=True, color="FF0070C0", italic=True)
            row += 1

        for k in range(len(process["product"])):
            product = process["product"][k]

            if i == 1:
                ws42[f"B{row}"] = product["production_class"]
                ws42[f"B{row}"].font = Font(name="Tahoma", size=10, bold=True, color="FF99004c", italic=True)
                row += 1

            for l in range(len(product["items"])):
                items = product["items"][l]
                text_out = items.get("type2_vehicle_outbound_display") or ""
                text_ret = items.get("type2_vehicle_outbound_display") or ""

                match_out = re.search(r'(\d+)%', text_out)
                match_ret = re.search(r'(\d+)%', text_ret)

                items["percent_vehicle_outbound"] = int(match_out.group(1)) if match_out else None
                items["percent_vehicle_return"] = int(match_ret.group(1)) if match_ret else None
                items["vehicle"] = re.split(r'\s+\d+%.*', text_out)[0].strip() if text_out else None

                ws42[f"B{row}"] = items["item_name"]
                ws42[f"C{row}"] = items["item_unit"]
                ws42[f"D{row}"] = items["item_fu_qty"]
                ws42[f"E{row}"] = items["distance"]
                ws42[f"F{row}"] = items["distance_source"]
                ws42[f"G{row}"] = items["type1_gas"]
                ws42[f"H{row}"] = items["type1_gas_unit"]
                ws42[f"I{row}"] = items["type1_gas_qty"]
                ws42[f"J{row}"] = items["type1_ef"]
                exec(check_source_form4_2_T1(items["type1_ef_source"], row, ws_name='ws42'))
                ws42[f"M{row}"] = items["type2_outbound_load"]
                ws42[f"N{row}"] = items["type2_return_load"]
                ws42[f"O{row}"] = items["vehicle"]
                ws42[f"P{row}"] = items["percent_vehicle_outbound"]
                ws42[f"Q{row}"] = items["percent_vehicle_return"]
                ws42[f"R{row}"] = items["type2_outbound_ef"]
                ws42[f"S{row}"] = items["type2_return_ef"]
                exec(check_source_form4_2_T2(items["type2_ef_source"], row, ws_name='ws42'))
                ws42[f"V{row}"] = items["type2_ef_source_ref"]
                ws42[f"W{row}"] = items["ratio"]
                proportion = (
                    (items.get("type2_outbound_load") or 0) * (items.get("type2_outbound_ef") or 0) +
                    (items.get("type2_return_load") or 0) * (items.get("type2_return_ef") or 0)
                )
                ws42[f"X{row}"] = round(float(proportion), 2)
                ws42[f"Y{row}"] = items["cut_off"]
                ws42[f"Z{row}"] = items["add_on_detail"]

                sum_proportion = sum_proportion + proportion
                row += 1  # ขยับแถวสำหรับ item ถัดไป

        # หลังจบแต่ละ process ให้เว้น 1 แถว
        row += 1

    # เติมสีสรุป process
    highlight_font = Font(name="Tahoma", size=10, bold=True, color="000000", italic=True)
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")  # พื้นหลังสีขาว
    thin_black = Side(style='thin', color='000000')  # เส้นดำบาง
    top_bottom_border = Border(top=thin_black, bottom=thin_black)

    for row_cells in ws42[f"B{row}:Z{row}"]:
        for cell in row_cells:
            cell.fill = fill_sum
    
    ws42[f"V{row}"] = "รวม"
    ws42[f"X{row}"] = sum_proportion
    for r in range(row, row + 1):
        for col in ["U", "V", "W", "X"]:
            cell = ws42[f"{col}{r}"]
            cell.font = highlight_font
            cell.fill = white_fill
            cell.border = top_bottom_border
    
    row += 1
    # Merge phase column และใส่ชื่อ phase
    ws42.merge_cells(f"A{phase_start_row}:A{row-1}")
    ws42[f"A{phase_start_row}"] = phase[i]

    r_start = row  # อัปเดต r_start สำหรับ phase ถัดไป

######## ---------------------------------------------------------------------

wb.save(output_xlsx)
print("Excel saved:", output_xlsx)

# ----------------- Export Sheet Fr-01 to PDF -----------------
def export_pdf_windows_excel(input_file, output_file):
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb_com = excel.Workbooks.Open(os.path.abspath(input_file))
    ws_com = wb_com.Sheets("Fr-04.2")
    ws_com.ExportAsFixedFormat(0, os.path.abspath(output_file))
    wb_com.Close(False)
    excel.Quit()

def export_pdf_libreoffice(input_file, output_file):
    if platform.system() == "Windows":
        libreoffice_path = r"C:\Program Files\LibreOffice\program\soffice.exe"
    else:
        libreoffice_path = "/usr/bin/libreoffice"
    if not os.path.exists(libreoffice_path):
        raise FileNotFoundError(f"LibreOffice not found at {libreoffice_path}")

    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)
    tmp_pdf_name = os.path.splitext(os.path.basename(input_file))[0] + ".pdf"
    tmp_pdf_path = os.path.join(output_dir, tmp_pdf_name)

    subprocess.run([
        libreoffice_path,
        "--headless",
        "--convert-to", "pdf",
        "--outdir", output_dir,
        input_file
    ], check=True)

    if os.path.exists(tmp_pdf_path):
        os.rename(tmp_pdf_path, output_file)
    else:
        raise FileNotFoundError(f"PDF not generated: {tmp_pdf_path}")
    return output_file

output_pdf = os.path.join(output_dir, f"{timestamp}_Company{company.get('company_id','')}_Product{product.get('product_id','')}_Fr042.pdf")

try:
    if platform.system() == "Windows":
        try:
            export_pdf_windows_excel(output_xlsx, output_pdf)
            print("PDF saved via Excel COM:", output_pdf)
        except Exception:
            export_pdf_libreoffice(output_xlsx, output_pdf)
            print("PDF saved via LibreOffice:", output_pdf)
    else:
        export_pdf_libreoffice(output_xlsx, output_pdf)
        print("PDF saved via LibreOffice:", output_pdf)
except Exception as e:
    print("PDF export failed:", e)