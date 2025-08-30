import os
import sys
import shutil
from datetime import datetime
import subprocess
import platform
from openpyxl.styles import PatternFill,Font,Alignment
from openpyxl.styles import Border, Side

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import requests
import json
import re

os.system("chcp 65001")  # เปลี่ยน code page เป็น UTF-8
sys.stdout.reconfigure(encoding='utf-8')  # Python 3.7+

# ----------------- Arguments -----------------
if len(sys.argv) < 3:
    print("Usage: python runExcel.py <company_name> <product_1>")
    sys.exit(1)
company_name = sys.argv[1]
product_1 = sys.argv[2]

# ----------------- API Call -----------------
url = f"http://localhost:5000/api/v1/f4-1/report/{product_1}"
resp = requests.get(url)
if resp.status_code != 200:
    print("เกิดข้อผิดพลาดในการเรียก API:", resp.status_code)
    sys.exit(1)

data = resp.json()

form41 = data.get("form41", {})
company = data.get("company", {})
product = data.get("product", {})
process = data.get("process", {})
report41Sum = data.get("report41Sum", [])[0]
finalproduct = data.get("finalproduct", [])[0]




# ----------------- Helper -----------------
def thai_date_format(iso_date_str):
    if not iso_date_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_date_str.replace("Z", "+00:00"))
    except ValueError:
        return "รูปแบบวันที่ไม่ถูกต้อง"

    months_th = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม",
                 "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม",
                 "พฤศจิกายน", "ธันวาคม"]
    return f"{dt.day} {months_th[dt.month]} {dt.year + 543}"

def check_source_form4_1(ef_source, row, ws_name='ws41'):
    ef_map = {
        'Self collect': 'H',
        'Supplier': 'I',
        'PCR Gen.': 'J',
        'TGO EF': 'K',
        'Int. DB': 'L',
        'Others': 'M'
    }

    if ef_source in ef_map:
        col = ef_map[ef_source]
        cell = f'{col}{row}'
        return f'{ws_name}["{cell}"] = "●"'
    else:
        return f'# ef_source \"{ef_source}\" not recognized'
    

# ----------------- Prepare Data -----------------
start = thai_date_format(product.get("collect_data_start"))
end = thai_date_format(product.get("collect_data_end"))
date_range = f"{start} - {end}"
submitted_date_thai = thai_date_format(product.get("submitted_date"))

techinfo_raw = product.get("product_techinfo") or "[]"
techinfo_list = json.loads(techinfo_raw) if techinfo_raw else []
product_techinfo_array = [str(i).strip() for i in techinfo_list]

# ----------------- File Paths -----------------
template_path = "ExcelReport/excel/form_CFP_pdf041.xlsx"
timestamp = datetime.now().strftime("%Y")
output_dir = "ExcelReport/output"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

output_xlsx = os.path.join(output_dir, f"{timestamp}_Company{company.get('company_id','')}_Product{product.get('product_id','')}.xlsx")
shutil.copy(template_path, output_xlsx)

# ----------------- Load Workbook -----------------

def set_borderThinTop(ws, cell_range,position):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            if position == 'left':
                cell.border = Border(left=Side(style='thin'))
            elif position == 'right':
                cell.border = Border(right=Side(style='thin'))
            elif position == 'top':
                cell.border = Border(
                    top=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )
            elif position == 'bottom':
                 cell.border = Border(
                    bottom=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )

######## ---------------------------------------------------------------------
wb = load_workbook(output_xlsx)
ws41 = wb["Fr-04.1"]
fill_sum = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid') 
r_start = 11
row = 11
# phase=["การได้มาของวัตถุดิบ","การผลิต","การกระจายสินค้า","การใช้งาน","การจัดการซาก"]
phase = ["การได้มาของวัตถุดิบ", "การผลิต", "การกระจายสินค้า", "การใช้งาน", "การจัดการซาก"]
columns = ["sum_lc1_emission", "sum_lc2_emission", "sum_lc3_emission", "sum_lc4_emission", "sum_lc5_emission"]
for i in range(len(phase)):
    phase_start_row = row  # เก็บแถวเริ่มต้นของ phase นี้
    FU = 0  # กำหนดค่าเริ่มต้นของ FU
    Qemission = 0  # กำหนดค่าเริ่มต้นของ GHG Emission
    sum_emission=0
    sum_lc_emission = report41Sum[columns[i]]
    for j in range(len(form41[i]["process"])):
        process = form41[i]["process"][j]
        process_start_row = row  # เก็บแถวเริ่มต้นของ process นี้

        if i < 2:
            ws41[f"B{row}"] = process["process_name"]
            ws41[f"B{row}"].font = Font(name="Tahoma", size=10, bold=True, color="FF0070C0", italic=True)
            row += 1
        
        for k in range(len(process["product"])):
            product = process["product"][k]

            if i == 1:
                ws41[f"B{row}"] = product["production_class"]
                ws41[f"B{row}"].font = Font(name="Tahoma", size=10, bold=True, color="FF99004c", italic=True)
                row += 1
            
            for l in range(len(product["items"])):
                items = product["items"][l]
                FU1 = round(float(items["item_quantity"]/finalproduct["output_quantity"]), 2) 

                ws41[f"B{row}"] = items["item_name"]
                ws41[f"C{row}"] = items["item_unit"]
                ws41[f"D{row}"] = items["item_quantity"]
                ws41[f"E{row}"] = FU1
                ws41[f"F{row}"] = items["lci_source_period"]
                ws41[f"G{row}"] = items["ef"]
                exec(check_source_form4_1(items["ef_source"], row, ws_name='ws41'))
                ws41[f"O{row}"] = items["ef_source_ref_display"]
                ws41[f"P{row}"] = items["ratio"]
                ef = round(float(items["ef"] if items["ef"] is not None else 0), 2)
                ws41[f"Q{row}"] = FU1 * ef
                ws41[f"R{row}"] = (FU1 * ef)/sum_lc_emission    
                ws41[f"S{row}"] = items["cut_off"]
                ws41[f"T{row}"] = items["description"]
                row += 1  # ขยับแถวสำหรับ item ถัดไป
                FU = FU + FU1
                Qemission= Qemission + (FU1 * ef)
                sum_emission=sum_emission+ ((FU1 * ef)/sum_lc_emission)


        # หลังจบแต่ละ process ให้เว้น 1 แถว
        row += 1
   

    # เติมสีสรุป process
    highlight_font = Font(name="Tahoma", size=10, bold=True, color="000000", italic=True)
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")  # พื้นหลังสีขาว
    thin_black = Side(style='thin', color='000000')  # เส้นดำบาง
    top_bottom_border = Border(top=thin_black, bottom=thin_black)
    for row_cells in ws41[f"B{row}:T{row}"]:
        for cell in row_cells:
            cell.fill = fill_sum
            cell.border = top_bottom_border
    ws41[f"C{row}"] = "รวม"
    ws41[f"E{row}"] = round(float(FU),2)
    ws41[f"Q{row}"] = round(float(Qemission),2)
    ws41[f"R{row}"] = round(float(sum_emission),2)

    for r in range(row, row + 1):
        for col in ["C", "D", "E", "Q", "R"]:
            cell = ws41[f"{col}{r}"]
            cell.font = highlight_font
            cell.fill = white_fill
            cell.border = top_bottom_border
    
    

    row += 1

    # Merge phase column และใส่ชื่อ phase
    ws41.merge_cells(f"A{phase_start_row}:A{row-1}")
    ws41[f"A{phase_start_row}"] = phase[i]

    r_start = row  # อัปเดต r_start สำหรับ phase ถัดไป

######## ---------------------------------------------------------------------

wb.save(output_xlsx)
print("Excel saved:", output_xlsx)

# ----------------- Export Sheet Fr-01 to PDF -----------------
def export_pdf_windows_excel(input_file, output_file):
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb_com = excel.Workbooks.Open(os.path.abspath(input_file))
    ws_com = wb_com.Sheets("Fr-04.1")
    ws_com.ExportAsFixedFormat(0, os.path.abspath(output_file))
    wb_com.Close(False)
    excel.Quit()

def export_pdf_libreoffice(input_file, output_file):
    if platform.system() == "Windows":
        libreoffice_path = r"C:\Program Files\LibreOffice\program\soffice.exe"
    else:
        libreoffice_path = "/usr/bin/libreoffice"
    if not os.path.exists(libreoffice_path):
        raise FileNotFoundError(f"LibreOffice not found at {libreoffice_path}")

    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)
    tmp_pdf_name = os.path.splitext(os.path.basename(input_file))[0] + ".pdf"
    tmp_pdf_path = os.path.join(output_dir, tmp_pdf_name)

    subprocess.run([
        libreoffice_path,
        "--headless",
        "--convert-to", "pdf",
        "--outdir", output_dir,
        input_file
    ], check=True)

    if os.path.exists(tmp_pdf_path):
        os.rename(tmp_pdf_path, output_file)
    else:
        raise FileNotFoundError(f"PDF not generated: {tmp_pdf_path}")
    return output_file

output_pdf = os.path.join(output_dir, f"{timestamp}_Company{company.get('company_id','')}_Product{product.get('product_id','')}_Fr041.pdf")

try:
    if platform.system() == "Windows":
        try:
            export_pdf_windows_excel(output_xlsx, output_pdf)
            print("PDF saved via Excel COM:", output_pdf)
        except Exception:
            export_pdf_libreoffice(output_xlsx, output_pdf)
            print("PDF saved via LibreOffice:", output_pdf)
    else:
        export_pdf_libreoffice(output_xlsx, output_pdf)
        print("PDF saved via LibreOffice:", output_pdf)
except Exception as e:
    print("PDF export failed:", e)