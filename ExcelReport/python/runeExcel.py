import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import units
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill,Font,Alignment
import requests
from datetime import datetime
import requests
import shutil
import json
import sys
from datetime import datetime
import io
import re
import os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

if len(sys.argv) < 3:
        print("Usage: python runExcel.py <company_name> <product_1>")
        sys.exit(1)

company_name = sys.argv[1]
product_1 = sys.argv[2]

url="http://localhost:5000/"
# url="http://178.128.123.212:5000/"
# company_name="1021"
# product_1="39"

form1 = url+"api/v1/f1/excel/"+company_name + "/" + product_1
form4_1= url+"api/v1/f4-1/report/"+product_1
form4_2= url+"api/v1/f4-2/form/"+product_1
form4_3= url+"api/v1/selfcollect/list/"+company_name + "/" + product_1
form6_1= url+"api/v1/f6-1/sum/" + product_1
form6_2= url+"api/v1/f6-2/sum/" + product_1
form5= url+"api/v1/f6-1/sumform4142/" + product_1


def thai_date_format(iso_date_str):
    if not iso_date_str:
        return ""  # หรือ return "ไม่ระบุวันที่" ตามต้องการ

    try:
        dt = datetime.fromisoformat(iso_date_str.replace("Z", "+00:00"))
    except ValueError:
        return "รูปแบบวันที่ไม่ถูกต้อง"

    months_th = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม",
                 "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม",
                 "พฤศจิกายน", "ธันวาคม"]

    day = dt.day
    month = months_th[dt.month]
    year = dt.year + 543

    return f"{day} {month} {year}"

# ///check Source Form4-1
def check_source_form4_1(ef_source, row, ws_name='ws41'):
    ef_map = {
        'Self collect': 'H',
        'Supplier': 'I',
        'PCR Gen.': 'J',
        'TGO EF': 'K',
        'Int. DB': 'L',
        'Others': 'M'
    }

    if ef_source in ef_map:
        col = ef_map[ef_source]
        cell = f'{col}{row}'
        return f'{ws_name}["{cell}"] = "●"'
    else:
        return f'# ef_source \"{ef_source}\" not recognized'

def check_source_form4_2_T1(type1_ef_source, row, ws_name='ws42'):
    ef_map = { 'TGO EF': 'K','Int. DB': 'L'}

    if type1_ef_source in ef_map:
        col = ef_map[type1_ef_source]
        cell = f'{col}{row}'
        return f'{ws_name}["{cell}"] = "●"'
    else:
        return f'# type1_ef_source \"{type1_ef_source}\" not recognized'

def check_source_form4_2_T2(type2_ef_source, row, ws_name='ws42'):
    ef_map = { 'TGO EF': 'T','Int. DB': 'U'}

    if type2_ef_source in ef_map:
        col = ef_map[type2_ef_source]
        cell = f'{col}{row}'
        return f'{ws_name}["{cell}"] = "●"'
    else:
        return f'# type2_ef_source \"{type2_ef_source}\" not recognized'

def check_source_form4_3(ef_source, row, ws_name='ws43'):
    ef_map = {
        'Self collect': 'I',
        'Supplier': 'J',
        'PCR Gen.': 'K',
        'TGO EF': 'L',
        'Int. DB': 'M',
        'Others': 'N'
    }

    if ef_source in ef_map:
        col = ef_map[ef_source]
        cell = f'{col}{row}'
        return f'{ws_name}["{cell}"] = "●"'
    else:
        return f'# ef_source \"{ef_source}\" not recognized'

def check_source_form4_3_T1(type1_ef_source, row, ws_name='ws43'):
    ef_map = { 'TGO EF': 'U','Int. DB': 'V'}

    if type1_ef_source in ef_map:
        col = ef_map[type1_ef_source]
        cell = f'{col}{row}'
        return f'{ws_name}["{cell}"] = "●"'
    else:
        return f'# type1_ef_source \"{type1_ef_source}\" not recognized'


def check_source_form4_3_T2(type2_ef_source, row, ws_name='ws42'):
    ef_map = { 'TGO EF': 'AE','Int. DB': 'AF'}

    if type2_ef_source in ef_map:
        col = ef_map[type2_ef_source]
        cell = f'{col}{row}'
        return f'{ws_name}["{cell}"] = "●"'
    else:
        return f'# type2_ef_source \"{type2_ef_source}\" not recognized'
        
# กำหนดเส้นกรอบแต่ละด้าน (เส้นบางและเส้นปะ)
thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

thin_top = Border(
    top=Side(border_style="thin", color="000000")
)

def set_borderThinTop(ws, cell_range,position):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            if position == 'left':
                cell.border = Border(left=Side(style='thin'))
            elif position == 'right':
                cell.border = Border(right=Side(style='thin'))
            elif position == 'top':
                cell.border = Border(
                    top=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )
            elif position == 'bottom':
                 cell.border = Border(
                    bottom=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )

    
def set_borderTop(ws, cell_range,position):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            if position == 'left':
                cell.border = Border(left=Side(style='dashed'))
            elif position == 'right':
                cell.border = Border(right=Side(style='dashed'))
            elif position == 'top':
                cell.border = Border(top=Side(style='dashed'))
            elif position == 'bottom':
                 cell.border = Border(
                    bottom=Side(style='dashed'),
                    left=Side(style='dashed'),
                    right=Side(style='dashed')
                )
            
def set_borderCenter(ws, cell_range, position):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            # อ่าน border เดิมก่อน
            border_old = cell.border

            # กำหนดด้านใหม่ตาม position
            new_border = Border(
                left=border_old.left if position != 'left' else Side(style='mediumDashed'),
                right=border_old.right if position != 'right' else Side(style='mediumDashed'),
                top=border_old.top if position != 'top' else Side(style='mediumDashed'),
                bottom=border_old.bottom if position != 'bottom' else Side(style='mediumDashed')
            )

            cell.border = new_border

def set_border(ws, cell_range):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = thin_border




# Fr-01, Fr.02, Fr.03
form1 = requests.get(form1)
if form1.status_code == 200:
    data = form1.json()
    productform1 = data.get("product", {})
    company = data.get("company", {})
    process = data.get("process", [])
    report41Sum = data.get("report41Sum", {})
    if not report41Sum:
        print("Warning: 'report41Sum' not found in API response.")
else:
    print("เกิดข้อผิดพลาดในการเรียก API:", form1.status_code)

   
start = thai_date_format(productform1["collect_data_start"])
end = thai_date_format(productform1["collect_data_end"])
date_range = f"{start} - {end}"
submitted_date_thai = thai_date_format(productform1.get("submitted_date"))
# techinfo_list = json.loads(product["product_techinfo"])
# product_techinfo_array = [item.strip() for item in techinfo_list]
techinfo_raw = productform1.get("product_techinfo") or "[]"
techinfo_list = json.loads(techinfo_raw) if techinfo_raw else []
product_techinfo_array = [str(i).strip() for i in techinfo_list]

form4_1 = requests.get(form4_1)
if form4_1.status_code == 200:
    data41 = form4_1.json()
    form41 = data41["form41"]
    company = data41["company"]
    product = data41["product"]
    process = data41["process"]
    report41Sum = data41["report41Sum"][0]
    finalproduct = data41["finalproduct"][0]
else:
    print("เกิดข้อผิดพลาดในการเรียก API:", form4_1.status_code)




response_form4_3 = requests.get(form4_3)
form43 = []  # กำหนดค่าเริ่มต้นก่อน
if response_form4_3.status_code == 200:
    data43 = response_form4_3.json()
    form43 = data43.get("processes", [])
else:
    print("เกิดข้อผิดพลาดในการเรียก API:", response_form4_3.status_code)
# ใช้งาน form43 ได้อย่างปลอดภัยหลังจากนี้
if len(form43) != 0:
    for process in form43:
        input_items = process.get("input", [])

response_form6_1 = requests.get(form6_1)
form61 = []
if response_form6_1.status_code == 200:
    data61 = response_form6_1.json()
    if isinstance(data61, list) and len(data61) > 0:
        form61 = data61[0]
    else:
        print("API ส่งข้อมูลว่างหรือไม่ได้อยู่ในรูปแบบ list")
else:
    print("เกิดข้อผิดพลาดในการเรียก API:", response_form6_1.status_code)

response_form6_2 = requests.get(form6_2)
form62 = []
if response_form6_2.status_code == 200:
    data62 = response_form6_2.json()
    if isinstance(data62, list) and len(data62) > 0:
        form62 = data62[0]
else:
    print("เกิดข้อผิดพลาดในการเรียก API:", response_form6_2.status_code)

response_form5 = requests.get(form5)
form5 = {}

if response_form5.status_code == 200:
    data5 = response_form5.json()
    if isinstance(data5, dict):
        form5 = data5
    elif isinstance(data5, list) and len(data5) > 0:
        form5 = data5[0]  # ถ้าอยากใช้ตัวแรกของ list
else:
    print("เกิดข้อผิดพลาดในการเรียก API:", response_form5.status_code)


file_path = "ExcelReport/excel/form_CFP.xlsx"
timestamp = datetime.now().strftime("%Y")
# product_id = productform1.get("product_id", "").replace(" ", "_")
output_path = f"ExcelReport/output/{timestamp}_Company{company['company_id']}_Product{productform1['product_id']}.xlsx"
shutil.copy(file_path, output_path)
process = data["process"]
wb = load_workbook(file_path)

fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid') 
fill_intput_head = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid')
fill_intput = PatternFill(start_color='FFCCFFCC', end_color='FFCCFFCC', fill_type='solid')
fill_output = PatternFill(start_color='FFFFCC99', end_color='FFFFCC99', fill_type='solid')
fill_waste = PatternFill(start_color='FFCCFFFF', end_color='FFCCFFFF', fill_type='solid')
fill_arrow = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
fill_tail_title = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
fill_light_blue = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')  # ฟ้าอ่อน
font_white = Font(color="FFFFFF", bold=True, size=10)
align_center = Alignment(horizontal="center", vertical="center")

######## ------------------Fr-01---------------------------------------------------
ws01 = wb["Fr-01"] # หรือระบุชื่อ sheet: wb["Sheet1"]

ws01["J2"] = date_range
ws01["F5"] = company.get("name", "")
ws01["F6"] = productform1.get("product_name_th", " ")
ws01["J10"] = productform1.get("product_name_th", " ")
ws01["J11"] = productform1.get("product_name_en", " ")
ws01["J12"] = productform1.get("scope", " ")
ws01["J13"] = str(productform1.get("FU_value", "")) + " " + productform1.get("FU_th_name", "") if productform1.get("FU_value") and productform1.get("FU_th_name") else "0"
ws01["J14"] = str(productform1.get("FU_value", "")) + " " + productform1.get("FU_en_name", "") if productform1.get("FU_value") and productform1.get("FU_en_name") else "0"
ws01["J15"] = str(productform1.get("PU_value", "")) + " " + productform1.get("PU_th_name", "") if productform1.get("PU_value") and productform1.get("PU_th_name") else "0"
ws01["J16"] = str(productform1.get("PU_value", "")) + " " +productform1.get("PU_en_name", "") if productform1.get("PU_value") and productform1.get("PU_en_name") else "0"
ws01["J17"] = productform1.get("sale_ratio", "")

start_row = 19
col = "I"
for idx, info in enumerate(product_techinfo_array):
    cell = f"{col}{start_row + idx}"
    ws01[cell] = info

ws01["J24"] = productform1.get("pcr_name", "") 
ws01["J25"] = submitted_date_thai

# image 
#image_url = product.get("product_photo", "").replace("public/", "Public/")
# image_url = product.get("product_photo", "")
# img = Image(image_url)  
# img.width = 300 
# img.height = 300  
# ws01.add_image(img, "B12")  

image_url = productform1.get("product_photo", "")

if image_url and os.path.isfile(image_url):
    img = Image(image_url)
    img.width = 300
    img.height = 300
    ws01.add_image(img, "B12")

######## ---------------------------------------------------------------------
ws02 = wb["Fr-02"]
for i in range(len(process)):
    input_names = []
    row = 13 + (3 * i)
    row_next = row + 1
    
    # ใส่ชื่อ process
    ws02.merge_cells(f"M{row}:O{row_next}")
    ws02[f"M{row}"] = process[i]['process_name']
    set_border(ws02, f"M{row}:O{row_next}")
    
    # ใส่คำว่า "ขนส่ง" และกำหนดเส้นขอบ
    ws02.merge_cells(f"I{row}:I{row_next}")
    ws02[f"I{row}"] = "ขนส่ง"
    set_border(ws02, f"I{row}:I{row_next}")
    ws02[f"H{row_next}"].border = thin_top
    ws02[f"J{row_next}"].border = thin_top
    ws02[f"K{row_next}"].border = thin_top
    
    
    # รวมชื่อ inputs ที่ไม่ใช่ "ผลิตภัณฑ์"
    input_names = [item['input_name'] for cat in process[i]['inputs'] for item in cat['items']]
    input_names_str = ', '.join(input_names)
    
    # ใส่ข้อมูลรวมชื่อ input
    ws02.merge_cells(f"C{row}:F{row_next}")
    ws02[f"C{row}"] = input_names_str
    set_border(ws02, f"C{row}:F{row_next}")

row_end= row_next + 2
# Box Input
ws02.merge_cells(f"B12:B{row_end-1}")
set_borderTop(ws02, f"B12:B{row_end-1}",'left')
ws02.merge_cells(f"G12:G{row_end-1}")
set_borderTop(ws02, f"G12:G{row_end-1}",'right')
ws02.merge_cells(f"B{row_end}:G{row_end}")
set_borderTop(ws02, f"B{row_end}:G{row_end}",'bottom')
# Box Process
ws02.merge_cells(f"L12:L{row_end-1}")
set_borderTop(ws02, f"L12:L{row_end-1}",'left')
ws02.merge_cells(f"P12:P{row_end-1}")
set_borderTop(ws02, f"P12:P{row_end-1}",'right')
ws02.merge_cells(f"L{row_end}:P{row_end}")
set_borderTop(ws02, f"L{row_end}:P{row_end}",'bottom')
# Box Center
# ws02.merge_cells(f"J8:J{row_end+3}")

set_borderCenter(ws02, f"J8:J{row_end+3}",'right')

# กำหนดค่าหลัก
ws02[f"A{row+10}"] = "จัดทำโดย"
ws02[f"F{row+10}"] = "เสร็จสิ้นวันที่"
ws02[f"N{row+10}"] = "วันที่แก้ไข"

# Merge E-F
ws02.merge_cells(f"F{row+10}:I{row+10}")
ws02[f"A{row+10}"].alignment = align_center
ws02[f"F{row+10}"].alignment = align_center

# ใส่ fill น้ำเงิน + font ขาว
for col in ["A", "F", "N"]:
    cell = ws02[f"{col}{row+10}"]
    cell.fill = fill_tail_title
    cell.font = font_white

# ใส่ฟ้าอ่อนใน B-D, G-I, K-M
for col_range in ["B:E", "J:M", "O:Q"]:
    start_col, end_col = col_range.split(":")
    cell_range = f"{start_col}{row+10}:{end_col}{row+10}"
    for row_cells in ws02[cell_range]:
        for cell in row_cells:
            cell.fill = fill_light_blue
    
######## ---------------------------------------------------------------------
ws03 = wb["Fr-03"]   
ws03["J1"] = date_range
ws03["D2"] = company.get("name", "")
ws03["D3"] = productform1.get("product_name_th", "")


row = 13 
for i in range(len(process)):
    input_names = []
    row_input = row
    row_input_offset = 0
    max_input_rows = 0 

    # ===== Input Section =====
    input_categories = process[i]['inputs']
    num_categories = len(input_categories)

    for j in range(max(1, num_categories)):
        if num_categories > 0:
            category = input_categories[j]
            items = category['items']
            num_items = len(items)
            input_title = category['input_title']
        else:
            items = []
            num_items = 0
            input_title = "ไม่มีข้อมูล"  # หรือจะเป็น "" ตามที่คุณต้องการ

        current_row = row_input + row_input_offset
        ws03.merge_cells(f"B{current_row}:D{current_row}")
        ws03[f"B{current_row}"] = input_title
        ws03[f"B{current_row}"].fill = fill_intput_head

        # ถ้ามี item จริง
        if num_items > 0:
            for k in range(num_items):
                ws03[f"B{current_row + k + 1}"] = items[k]['input_name']
                ws03[f"C{current_row + k + 1}"] = items[k]['input_quantity']
                ws03[f"D{current_row + k + 1}"] = items[k]['input_unit']

                ws03[f"B{current_row + k + 1}"].fill = fill_intput
                ws03[f"C{current_row + k + 1}"].fill = fill_intput
                ws03[f"D{current_row + k + 1}"].fill = fill_intput
        else:
            # ใส่ช่องว่างแถวล่างไว้ 1 แถวพร้อมสี
            ws03[f"B{current_row + 1}"] = ""
            ws03[f"C{current_row + 1}"] = ""
            ws03[f"D{current_row + 1}"] = ""

            ws03[f"B{current_row + 1}"].fill = fill_intput
            ws03[f"C{current_row + 1}"].fill = fill_intput
            ws03[f"D{current_row + 1}"].fill = fill_intput

        row_input_offset += 1 + max(1, num_items)
        max_input_rows += 1 + max(1, num_items)
    # ===== Process Block ด้านขวา =====
    process_row = row
    ws03.merge_cells(f"F{process_row}:I{process_row}")
    ws03[f"F{process_row}"] = process[i]['ordering']
    ws03[f"F{process_row}"].fill = fill
    set_borderThinTop(ws03, f"F{process_row}:I{process_row}", 'top')

    ws03.merge_cells(f"F{process_row + 1}:I{process_row + 5}")
    ws03[f"F{process_row + 1}"] = process[i]['process_name']
    ws03[f"F{process_row + 1}"].fill = fill
    set_borderThinTop(ws03, f"F{process_row + 1}:I{process_row + 5}", 'bottom')

    # ===== output =====
    ws03[f"I{process_row + 7}"] = "ผลิตภัณฑ์"
    ws03[f"I{process_row + 7}"].fill = fill_intput_head 

    for j in range(len(process[i]['outputs'])):
        output = process[i]['outputs'][j]
        ws03[f"I{process_row + 8 + j}"] = output['output_name']
        ws03[f"J{process_row + 8 + j}"] = output['output_quantity']
        ws03[f"K{process_row + 8 + j}"] = output['output_unit']
        ws03[f"I{process_row + 8 + j}"].fill = fill_output
        ws03[f"J{process_row + 8 + j}"].fill = fill_output
        ws03[f"K{process_row + 8 + j}"].fill = fill_output

    # ===== waste =====
    waste_head=["ผลิตภัณฑ์ร่วม", "ของเสีย"]
    row_input_offset = 0
    current_row = row_input + row_input_offset
    waste_total_rows = 0 

    if process[i]['wastes'] == []:
        ws03.merge_cells(f"K{current_row}:M{current_row}")
        ws03[f"K{current_row}"] = "ผลิตภัณฑ์ร่วม"
        ws03[f"K{current_row}"].fill = fill_intput_head
        ws03.merge_cells(f"K{current_row + 1}:M{current_row + 1}")
        ws03[f"K{current_row + 1}"].fill = fill_waste

        ws03.merge_cells(f"K{current_row+2}:M{current_row+2}")
        ws03[f"K{current_row+2}"] = "ของเสีย"
        ws03[f"K{current_row+2}"].fill = fill_intput_head
        ws03.merge_cells(f"K{current_row + 3}:M{current_row + 3}")
        ws03[f"K{current_row + 3}"].fill = fill_waste
    else:
        len_waste = len(process[i]['wastes'])
        if(len_waste == 1):
            if process[i]['wastes'][0]['waste_cat_name'] == waste_head[0]:
                ws03.merge_cells(f"K{current_row}:M{current_row}")
                ws03[f"K{current_row}"] = waste_head[0]
                ws03[f"K{current_row}"].fill = fill_intput_head
                for j in range(len(process[i]['wastes'][0]['items'])):
                    item = process[i]['wastes'][0]['items'][j]
                    ws03[f"K{current_row + 2 + j}"] = item['waste_name']
                    ws03[f"L{current_row + 2 + j}"] = item['waste_qty']
                    ws03[f"M{current_row + 2 + j}"] = item['waste_unit']
                    ws03[f"K{current_row + 2 + j}"].fill = fill_waste
                    ws03[f"L{current_row + 2 + j}"].fill = fill_waste
                    ws03[f"M{current_row + 2 + j}"].fill = fill_waste
                
                
                waste_total_rows += len(process[i]['wastes'][0]['items']) + 2

                ws03.merge_cells(f"K{waste_total_rows}:M{waste_total_rows}")
                ws03[f"K{waste_total_rows}"] = waste_head[0]
                ws03[f"K{waste_total_rows}"].fill = fill_intput_head
                ws03.merge_cells(f"K{waste_total_rows+1}:M{waste_total_rows+1}")
                ws03[f"K{waste_total_rows+1}"].fill = fill_waste

            else:

                ws03.merge_cells(f"K{current_row}:M{current_row}")
                ws03[f"K{current_row}"] = waste_head[0]
                ws03[f"K{current_row}"].fill = fill_intput_head
                ws03.merge_cells(f"K{current_row + 1}:M{current_row + 1}")
                ws03[f"K{current_row + 1}"].fill = fill_waste
                ws03.merge_cells(f"K{current_row + 2}:M{current_row + 2}")
                ws03[f"K{current_row + 2}"] = waste_head[1] 
                ws03[f"K{current_row + 2}"].fill = fill_intput_head


                for j in range(len(process[i]['wastes'][0]['items'])):
                    item = process[i]['wastes'][0]['items'][j]
                    ws03[f"K{current_row + 3 + j}"] = item['waste_name']
                    ws03[f"L{current_row + 3 + j}"] = item['waste_qty']
                    ws03[f"M{current_row + 3 + j}"] = item['waste_unit']
                    ws03[f"K{current_row + 3 + j}"].fill = fill_waste
                    ws03[f"L{current_row + 3 + j}"].fill = fill_waste
                    ws03[f"M{current_row + 3 + j}"].fill = fill_waste

                waste_total_rows += len(process[i]['wastes'][0]['items']) + 2

        for j in range(len(process[i]['wastes'])):

            if process[i]['wastes'][0]['waste_cat_name']==waste_head[j]:
                ws03.merge_cells(f"K{current_row}:M{current_row}")
                ws03[f"K{current_row}"] = waste_head[0]
                ws03[f"K{current_row}"].fill = fill_intput_head
                 # เขียนรายการผลิตภัณฑ์ร่วม
                for j in range(len(process[i]['wastes'][0]['items'])):
                    item = process[i]['wastes'][0]['items'][j]
                    ws03[f"K{current_row + 2 + j}"] = item['waste_name']
                    ws03[f"L{current_row + 2 + j}"] = item['waste_qty']
                    ws03[f"M{current_row + 2 + j}"] = item['waste_unit']
                    ws03[f"K{current_row + 2 + j}"].fill = fill_waste
                    ws03[f"L{current_row + 2 + j}"].fill = fill_waste
                    ws03[f"M{current_row + 2 + j}"].fill = fill_waste

                waste_total_rows += len(process[i]['wastes'][0]['items']) + 2
            else:
                ws03.merge_cells(f"K{current_row}:M{current_row}")
                ws03[f"K{current_row}"] = waste_head[0]
                ws03[f"K{current_row}"].fill = fill_intput_head
                ws03.merge_cells(f"K{current_row + 1}:M{current_row + 1}")
            
            ws03[f"K{current_row + 1}"].fill = fill_waste

    # ===== ปรับ row สำหรับ process ถัดไป =====
    ws03[f"E{process_row+2}"].fill = fill_arrow
    ws03[f"J{process_row+2}"].fill = fill_arrow
    # ws03.merge_cells(f"H{process_row+6}:H{process_row+6+4}")
    # ws03[f"H{process_row+6}"].fill = fill_arrow

    if (max_input_rows <=6):
        u=process_row+6
        l=process_row+9
    else:
        x=max_input_rows-9
        u=process_row+6
        l=process_row+10+x
    
    if j< len(process)-1:
        ws03.merge_cells(f"G{u}:G{l}")
        ws03[f"G{u}"].fill = fill_arrow

    row += max(max_input_rows + 2, 10) 


# กำหนดค่าหลัก
ws03[f"A{row+5}"] = "จัดทำโดย"
ws03[f"E{row+5}"] = "เสร็จสิ้นวันที่"
ws03[f"J{row+5}"] = "วันที่แก้ไข"

# Merge E-F
ws03.merge_cells(f"E{row+5}:F{row+5}")
ws03[f"E{row+5}"].alignment = align_center

# ใส่ fill น้ำเงิน + font ขาว
for col in ["A", "E", "J"]:
    cell = ws03[f"{col}{row+5}"]
    cell.fill = fill_tail_title
    cell.font = font_white

# ใส่ฟ้าอ่อนใน B-D, G-I, K-M
for col_range in ["B:D", "G:I", "K:M"]:
    start_col, end_col = col_range.split(":")
    cell_range = f"{start_col}{row+5}:{end_col}{row+5}"
    for row_cells in ws03[cell_range]:
        for cell in row_cells:
            cell.fill = fill_light_blue
######## ---------------------------------------------------------------------

ws41 = wb["Fr-04.1 "]
fill_sum = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid') 
r_start = 11
row = 11
# phase=["การได้มาของวัตถุดิบ","การผลิต","การกระจายสินค้า","การใช้งาน","การจัดการซาก"]
phase = ["การได้มาของวัตถุดิบ", "การผลิต", "การกระจายสินค้า", "การใช้งาน", "การจัดการซาก"]
columns = ["sum_lc1_emission", "sum_lc2_emission", "sum_lc3_emission", "sum_lc4_emission", "sum_lc5_emission"]
for i in range(len(phase)):
    phase_start_row = row  # เก็บแถวเริ่มต้นของ phase นี้
    FU = 0  # กำหนดค่าเริ่มต้นของ FU
    Qemission = 0  # กำหนดค่าเริ่มต้นของ GHG Emission
    sum_emission=0
    sum_lc_emission = report41Sum.get(columns[i], 1)
    for j in range(len(form41[i]["process"])):
        process = form41[i]["process"][j]
        process_start_row = row  # เก็บแถวเริ่มต้นของ process นี้

        if i < 2:
            ws41[f"B{row}"] = process["process_name"]
            ws41[f"B{row}"].font = Font(name="Tahoma", size=10, bold=True, color="FF0070C0", italic=True)
            row += 1
        
        for k in range(len(process["product"])):
            product = process["product"][k]

            if i == 1:
                ws41[f"B{row}"] = product["production_class"]
                ws41[f"B{row}"].font = Font(name="Tahoma", size=10, bold=True, color="FF99004c", italic=True)
                row += 1
            
            for l in range(len(product["items"])):
                items = product["items"][l]
                FU1 = round(float(items["item_quantity"]/finalproduct["output_quantity"]), 2) 

                ws41[f"B{row}"] = items["item_name"]
                ws41[f"C{row}"] = items["item_unit"]
                ws41[f"D{row}"] = items["item_quantity"]
                ws41[f"E{row}"] = FU1
                ws41[f"F{row}"] = items["lci_source_period"]
                ws41[f"G{row}"] = items["ef"]
                exec(check_source_form4_1(items["ef_source"], row, ws_name='ws41'))
                ws41[f"O{row}"] = items["ef_source_ref_display"]
                ws41[f"P{row}"] = items["ratio"]
                ef = round(float(items["ef"] if items["ef"] is not None else 0), 2)
                ws41[f"Q{row}"] = FU1 * ef
                ws41[f"R{row}"] = (FU1 * ef) / sum_lc_emission if sum_lc_emission else 0    
                ws41[f"S{row}"] = items["cut_off"]
                ws41[f"T{row}"] = items["description"]
                row += 1  # ขยับแถวสำหรับ item ถัดไป
                FU = FU + FU1
                Qemission= Qemission + (FU1 * ef)
                sum_emission += (FU1 * ef) / sum_lc_emission if sum_lc_emission else 0


        # หลังจบแต่ละ process ให้เว้น 1 แถว
        row += 1
   

    # เติมสีสรุป process
    highlight_font = Font(name="Tahoma", size=10, bold=True, color="000000", italic=True)
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")  # พื้นหลังสีขาว
    thin_black = Side(style='thin', color='000000')  # เส้นดำบาง
    top_bottom_border = Border(top=thin_black, bottom=thin_black)
    for row_cells in ws41[f"B{row}:T{row}"]:
        for cell in row_cells:
            cell.fill = fill_sum
            cell.border = top_bottom_border
    ws41[f"C{row}"] = "รวม"
    ws41[f"E{row}"] = round(float(FU),2)
    ws41[f"Q{row}"] = round(float(Qemission),2)
    ws41[f"R{row}"] = round(float(sum_emission),2)

    for r in range(row, row + 1):
        for col in ["C", "D", "E", "Q", "R"]:
            cell = ws41[f"{col}{r}"]
            cell.font = highlight_font
            cell.fill = white_fill
            cell.border = top_bottom_border
    
    

    row += 1

    # Merge phase column และใส่ชื่อ phase
    ws41.merge_cells(f"A{phase_start_row}:A{row-1}")
    ws41[f"A{phase_start_row}"] = phase[i]

    r_start = row  # อัปเดต r_start สำหรับ phase ถัดไป


######## ---------------------------------------------------------------------

form4_2 = requests.get(form4_2)
if form4_2.status_code == 200:
    data42 = form4_2.json()
    form42 = data42["form42"]
    company = data42["company"]
    product = data42["product"]
    process = data42["process"]
    report42Sum = data42["report42Sum"][0]
else:
    print("เกิดข้อผิดพลาดในการเรียก API:", form4_2.status_code)


ws42 = wb["Fr-04.2 "]
fill_sum = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid') 
r_start = 12
row = 12
phase = ["การได้มาของวัตถุดิบ", "การผลิต", "การกระจายสินค้า", "การใช้งาน", "การจัดการซาก"]
for i in range(len(phase)):
    phase_start_row = row  # เก็บแถวเริ่มต้นของ phase นี้
    sum_proportion = 0  # กำหนดค่าเริ่มต้นของ sum_proportion

    for j in range(len(form42[i]["process"])):
        process = form42[i]["process"][j]
        process_start_row = row  # เก็บแถวเริ่มต้นของ process นี้

        if i < 2:
            ws42[f"B{row}"] = process["process_name"]
            ws42[f"B{row}"].font = Font(name="Tahoma", size=10, bold=True, color="FF0070C0", italic=True)
            row += 1

        for k in range(len(process["product"])):
            product = process["product"][k]

            if i == 1:
                ws42[f"B{row}"] = product["production_class"]
                ws42[f"B{row}"].font = Font(name="Tahoma", size=10, bold=True, color="FF99004c", italic=True)
                row += 1

            for l in range(len(product["items"])):
                items = product["items"][l]
                text_out = items.get("type2_vehicle_outbound_display") or ""
                text_ret = items.get("type2_vehicle_outbound_display") or ""

                match_out = re.search(r'(\d+)%', text_out)
                match_ret = re.search(r'(\d+)%', text_ret)

                items["percent_vehicle_outbound"] = int(match_out.group(1)) if match_out else None
                items["percent_vehicle_return"] = int(match_ret.group(1)) if match_ret else None
                items["vehicle"] = re.split(r'\s+\d+%.*', text_out)[0].strip() if text_out else None

                ws42[f"B{row}"] = items["item_name"]
                ws42[f"C{row}"] = items["item_unit"]
                ws42[f"D{row}"] = items["item_fu_qty"]
                ws42[f"E{row}"] = items["distance"]
                ws42[f"F{row}"] = items["distance_source"]
                ws42[f"G{row}"] = items["type1_gas"]
                ws42[f"H{row}"] = items["type1_gas_unit"]
                ws42[f"I{row}"] = items["type1_gas_qty"]
                ws42[f"J{row}"] = items["type1_ef"]
                exec(check_source_form4_2_T1(items["type1_ef_source"], row, ws_name='ws42'))
                ws42[f"M{row}"] = items["type2_outbound_load"]
                ws42[f"N{row}"] = items["type2_return_load"]
                ws42[f"O{row}"] = items["vehicle"]
                ws42[f"P{row}"] = items["percent_vehicle_outbound"]
                ws42[f"Q{row}"] = items["percent_vehicle_return"]
                ws42[f"R{row}"] = items["type2_outbound_ef"]
                ws42[f"S{row}"] = items["type2_return_ef"]
                exec(check_source_form4_2_T2(items["type2_ef_source"], row, ws_name='ws42'))
                ws42[f"V{row}"] = items["type2_ef_source_ref"]
                ws42[f"W{row}"] = items["ratio"]
                proportion = (
                    (items.get("type2_outbound_load") or 0) * (items.get("type2_outbound_ef") or 0) +
                    (items.get("type2_return_load") or 0) * (items.get("type2_return_ef") or 0)
                )
                ws42[f"X{row}"] = round(float(proportion), 2)
                ws42[f"Y{row}"] = items["cut_off"]
                ws42[f"Z{row}"] = items["add_on_detail"]

                sum_proportion = sum_proportion + proportion
                row += 1  # ขยับแถวสำหรับ item ถัดไป

        # หลังจบแต่ละ process ให้เว้น 1 แถว
        row += 1

    # เติมสีสรุป process
    highlight_font = Font(name="Tahoma", size=10, bold=True, color="000000", italic=True)
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")  # พื้นหลังสีขาว
    thin_black = Side(style='thin', color='000000')  # เส้นดำบาง
    top_bottom_border = Border(top=thin_black, bottom=thin_black)

    for row_cells in ws42[f"B{row}:Z{row}"]:
        for cell in row_cells:
            cell.fill = fill_sum
    
    ws42[f"V{row}"] = "รวม"
    ws42[f"X{row}"] = sum_proportion
    for r in range(row, row + 1):
        for col in ["U", "V", "W", "X"]:
            cell = ws42[f"{col}{r}"]
            cell.font = highlight_font
            cell.fill = white_fill
            cell.border = top_bottom_border
    
    row += 1
    # Merge phase column และใส่ชื่อ phase
    ws42.merge_cells(f"A{phase_start_row}:A{row-1}")
    ws42[f"A{phase_start_row}"] = phase[i]

    r_start = row  # อัปเดต r_start สำหรับ phase ถัดไป

######## ----Fr-04.3-----------------------------------------------------------

ws43 = wb["Fr-04.3"]
fill_sum = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid') 
row=11
r_start = 11


if form43:
    for k in range(len(form43)):
        process = form43[k]
        input_items = process["input"]
        output_items = process["output"]
        name = process["self_collect_name"]

        row = r_start  # เก็บตำแหน่งเริ่มต้นสำหรับ merge cell
        clean_name = name.replace("Fr04.3 ", "", 1)

        ws43[f"B{r_start }"] = "Input"
        ws43[f"B{r_start }"].font = Font(name="Tahoma", size=10, bold=True, color="FF0070C0", italic=True)

        # ----- INPUT -----
        for i, input_item in enumerate(input_items):
            row_num = r_start + i + 1
            ws43[f"B{row_num}"] = input_item.get("item_name", "")
            ws43[f"C{row_num}"] = input_item.get("item_unit", "")
            ws43[f"D{row_num}"] = input_item.get("item_qty", "")
            ws43[f"E{row_num}"] = input_item.get("item_fu_qty", "")
            ws43[f"F{row_num}"] = input_item.get("item_source", "") 
            ws43[f"G{row_num}"] = input_item.get("item_ef", "") 
            exec(check_source_form4_1(input_item.get("item_ef_source", ""), row_num, ws_name='ws43'))
            ws43[f"O{row_num}"] = input_item.get("item_ef_source_ref", "")
            ws43[f"P{row_num}"] = input_item.get("item_emission", "")

            ws43[f"Q{row_num}"] = input_item.get("type1_gas", "")
            ws43[f"R{row_num}"] = input_item.get("type1_gas_unit", "")
            ws43[f"S{row_num}"] = input_item.get("type1_gas_qty", "")
            ws43[f"T{row_num}"] = input_item.get("type1_ef", "")
            exec(check_source_form4_3_T1(input_item.get("type1_ef_source", ""), row_num, ws_name='ws43'))

            ws43[f"W{row_num}"] = input_item.get("type2_distance", "")
            ws43[f"X{row_num}"] = input_item.get("type2_outbound_load", "")
            ws43[f"Y{row_num}"] = input_item.get("type2_return_load", "")
            ws43[f"Z{row_num}"] = input_item.get("type2_vehicle", "")
            ws43[f"AA{row_num}"] = input_item.get("type2_outbound_load_percent", "")
            ws43[f"AB{row_num}"] = input_item.get("type2_return_load_percent", "")
            ws43[f"AC{row_num}"] = input_item.get("type2_outbound_ef", "")
            ws43[f"AD{row_num}"] = input_item.get("type2_return_ef", "")
            exec(check_source_form4_3_T2(input_item.get("type2_ef_source", ""), row_num, ws_name='ws43'))
            ws43[f"AG{row_num}"] = input_item.get("type2_ef_source_ref", "")
            ws43[f"AH{row_num}"] = input_item.get("transport_emission", "")
            ws43[f"AI{row_num}"] = input_item.get("total_emission", "")
            ws43[f"AJ{row_num}"] = input_item.get("ratio", "")
            ws43[f"AK{row_num}"] = input_item.get("cut_off", "")
            ws43[f"AL{row_num}"] = input_item.get("add_on_detail", "")

        r_start += len(input_items) + 1

        # ----- OUTPUT -----
        ws43[f"B{r_start}"] = "Output"
        ws43[f"B{r_start}"].font = Font(name="Tahoma", size=10, bold=True, color="FF0070C0", italic=True)

        for i, output_item in enumerate(output_items):
            row_num = r_start + i + 1
            ws43[f"B{row_num}"] = output_item.get("item_name", "")
            ws43[f"C{row_num}"] = output_item.get("item_unit", "")
            ws43[f"D{row_num}"] = round(float(output_item.get("item_qty", "")), 2)
            ws43[f"E{row_num}"] = round(float(output_item.get("item_fu_qty", "")), 2)
            ws43[f"F{row_num}"] = output_item.get("item_source", "")
            ws43[f"G{row_num}"] = output_item.get("item_ef", "")
            exec(check_source_form4_1(output_item.get("item_ef_source", ""), row_num, ws_name='ws43'))
            ws43[f"O{row_num}"] = output_item.get("item_ef_source_ref", "")
            ws43[f"P{row_num}"] = output_item.get("item_emission", "")

            ws43[f"Q{row_num}"] = output_item.get("type1_gas", "")
            ws43[f"R{row_num}"] = output_item.get("type1_gas_unit", "")
            ws43[f"S{row_num}"] = output_item.get("type1_gas_qty", "")
            ws43[f"T{row_num}"] = output_item.get("type1_ef", "")
            exec(check_source_form4_3_T1(output_item.get("type1_ef_source", ""), row_num, ws_name='ws43'))

            ws43[f"W{row_num}"] = output_item.get("type2_distance", "")
            ws43[f"X{row_num}"] = output_item.get("type2_outbound_load", "")
            ws43[f"Y{row_num}"] = output_item.get("type2_return_load", "")
            ws43[f"Z{row_num}"] = output_item.get("type2_vehicle", "")
            ws43[f"AA{row_num}"] = output_item.get("type2_outbound_load_percent", "")
            ws43[f"AB{row_num}"] = output_item.get("type2_return_load_percent", "")
            ws43[f"AC{row_num}"] = output_item.get("type2_outbound_ef", "")
            ws43[f"AD{row_num}"] = output_item.get("type2_return_ef", "")
            exec(check_source_form4_3_T2(output_item.get("type2_ef_source", ""), row_num, ws_name='ws43'))
            ws43[f"AG{row_num}"] = output_item.get("type2_ef_source_ref", "")
            ws43[f"AH{row_num}"] = output_item.get("transport_emission", "")
            ws43[f"AI{row_num}"] = output_item.get("total_emission", "")
            ws43[f"AJ{row_num}"] = output_item.get("ratio", "")
            ws43[f"AK{row_num}"] = output_item.get("cut_off", "")
            ws43[f"AL{row_num}"] = output_item.get("add_on_detail", "")

        last_row = r_start + len(output_items) + 2
        ws43.merge_cells(f"A{row}:A{last_row}")
        ws43[f"A{row}"] = clean_name

        # เติมสีให้ summary row
        for row_cells in ws43[f"B{last_row}:AL{last_row}"]:
            for cell in row_cells:
                cell.fill = fill_sum

        r_start = last_row + 1  # เตรียมรอบถัดไป

######## ----Fr-05-----------------------------------------------------------
ws5 = wb["Fr-05"]
ws5["C14"] = round(float(report41Sum["sum_lc1_emission"]), 2) if report41Sum["sum_lc1_emission"] is not None else 0.00
ws5["C15"] = round(float(report41Sum["sum_lc2_emission"]), 2) if report41Sum["sum_lc2_emission"] is not None else 0.00
ws5["C16"] = round(float(report41Sum["sum_lc3_emission"]), 2) if report41Sum["sum_lc3_emission"] is not None else 0.00
ws5["C17"] = round(float(report41Sum["sum_lc4_emission"]), 2) if report41Sum["sum_lc4_emission"] is not None else 0.00
ws5["C18"] = round(float(report41Sum["sum_lc5_emission"]), 2) if report41Sum["sum_lc5_emission"] is not None else 0.00

ws5[f"D14"] =  round(float(report42Sum["lc1_transport_emission"]), 2) if report42Sum["lc1_transport_emission"] is not None else 0.00
ws5[f"D15"] =  round(float(report42Sum["lc2_transport_emission"]), 2) if report42Sum["lc2_transport_emission"] is not None else 0.00 
ws5[f"D16"] =  round(float(report42Sum["lc3_transport_emission"]), 2) if report42Sum["lc3_transport_emission"] is not None else 0.00
ws5[f"D17"] =  round(float(report42Sum["lc4_transport_emission"]), 2) if report42Sum["lc4_transport_emission"] is not None else 0.00
ws5[f"D18"] =  round(float(report42Sum["lc5_transport_emission"]), 2) if report42Sum["lc5_transport_emission"] is not None else 0.00





######## ----Fr-06.1-----------------------------------------------------------

ws61 = wb["Fr-06.1"]
if form61:
    ws61[f"C13"] =  round(float(form61["lc1_based_emission"]), 2)
    ws61[f"C14"] =  round(float(form61["lc2_based_emission"]), 2)
    ws61[f"C15"] =  round(float(form61["lc3_based_emission"]), 2)
    ws61[f"C16"] =  round(float(form61["lc4_based_emission"]), 2)
    ws61[f"C17"] =  round(float(form61["lc5_based_emission"]), 2)
    ws61[f"C18"] =  round(float(form61["land_used_based_emission"]), 2)

    ws61[f"D13"] = round(float(form5["sum_lc1_emission"]), 2)
    ws61[f"D14"] = round(float(form5["sum_lc2_emission"]), 2)
    ws61[f"D15"] = round(float(form5["sum_lc3_emission"]), 2)
    ws61[f"D16"] = round(float(form5["sum_lc4_emission"]), 2)
    ws61[f"D17"] = round(float(form5["sum_lc5_emission"]), 2)

######## ----Fr-06.2-----------------------------------------------------------

ws62 = wb["Fr-06.2"]
start_row = 11
if form62:
    for i, form62 in enumerate(data62):
        row = start_row + i
        ws62[f"B{row}"] = round(float(form62["std_emission"]), 2)
        ws62[f"C{row}"] = round(float(form62["product_emission"]), 2)
        ws62[f"D{row}"] = round(float(form62["diff_emission"]), 2)
        ws62[f"E{row}"] = form62["std_emission_source"] or ""

wb.save(output_path)
print(output_path)